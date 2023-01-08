from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QDialog, QApplication, QMainWindow
from PyQt5.QtWidgets import QFileDialog
from PyQt5 import uic
import sys
import os
import subprocess
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from func.pharmacy_scrapper import PharmacyScrapper
from func.helpers import add_product_frame, remove_product_frame

class MainWindow(QMainWindow):
	def __init__(self):
		super(MainWindow, self).__init__()

		# Load the ui file
		uic.loadUi(os.path.join('ui', 'main.ui'), self) 

		# Initial value: list of pharmacies
		self.pharmacies = ['Aptelia', 'Cefarm24', 'Gemini', 'Melissa', 'Ziko']

		# -------- App widgets settings --------
		# Add the 'turn_off_the_pharmacy' function to pharmacy buttons
		self.pharmacy_button_1.clicked.connect(lambda: self.turn_off_the_pharmacy(self.pharmacy_button_1))
		self.pharmacy_button_2.clicked.connect(lambda: self.turn_off_the_pharmacy(self.pharmacy_button_2))
		self.pharmacy_button_3.clicked.connect(lambda: self.turn_off_the_pharmacy(self.pharmacy_button_3))
		self.pharmacy_button_4.clicked.connect(lambda: self.turn_off_the_pharmacy(self.pharmacy_button_4))
		self.pharmacy_button_5.clicked.connect(lambda: self.turn_off_the_pharmacy(self.pharmacy_button_5))

		# Add the 'add_product_frame' function to the 'add_product_pushButton'
		self.add_product_pushButton.clicked.connect(lambda: add_product_frame(self.scrollAreaWidgetLayout, self.scrollAreaWidget))

		# Add the 'remove_product_frame' function to the 'remove_prodcut_pushButton'
		self.remove_product_pushButton.clicked.connect(lambda: remove_product_frame(self.scrollAreaWidgetLayout))

		# Add the 'load_products' function to the 'run_pushButton'
		self.run_pushButton.clicked.connect(lambda: self.load_products())

	def turn_off_the_pharmacy(self, button):
		# If the button with the name of the pharmacy is checked, remove the name of the pharmacy from the self.pharmacies list
		if button.isChecked():
			self.pharmacies.remove(button.text())
		else:
			self.pharmacies.append(button.text())
		# Sort the self.pharmacies list
		self.pharmacies.sort()
	
	def load_products(self):
		# Initial value: empty products list
		self.products = []
		# Find each QtWidgets.QTextBrowser widget and assign its text value to the product list
		for i in range(len(self.findChildren(QtWidgets.QTextBrowser))):
			self.products.append(self.findChildren(QtWidgets.QTextBrowser)[i].toPlainText())
		# Run check_prices function
		self.check_prices()

	def check_prices(self):
		# Initial value: empty results list
		self.results = []
		# Check price from pharmacies from the self.pharmacies list 
		for product in self.products:
			product_result = []
			if "Aptelia" in self.pharmacies:
				# Check price from Aptelia pharmacy
					scrapper = PharmacyScrapper(product)
					dict_1 = scrapper.check_aptelia_pharmacy()
					product_result.append(dict_1)
			if "Cefarm24" in self.pharmacies:
				# Check price from Cefarm24 pharmacy
					scrapper = PharmacyScrapper(product)
					dict_2 = scrapper.check_cefarm_pharmacy()
					product_result.append(dict_2)
			if "Gemini" in self.pharmacies:
				# Check price from Gemini pharmacy
					scrapper = PharmacyScrapper(product)
					dict_3 = scrapper.check_gemini_pharmacy()
					product_result.append(dict_3)
			if "Melissa" in self.pharmacies:
				# Check price from Melissa pharmacy
					scrapper = PharmacyScrapper(product)
					dict_4 = scrapper.check_melissa_pharmacy()
					product_result.append(dict_4)
			if "Ziko" in self.pharmacies:
				# Check price from Ziko pharmacy
					scrapper = PharmacyScrapper(product)
					dict_5 = scrapper.check_ziko_pharmacy()
					product_result.append(dict_5)	
			# Add results to results list
			self.results.append(product_result)
		# Open popup window
		self.open_popup_window()		

	def open_popup_window(self):
		self.popup_window = PopupWindow(self.products, self.results)
		self.popup_window.show()
		self.popup_window.raise_()
		self.popup_window.activateWindow()
	

class PopupWindow(QMainWindow):
	def __init__(self, products, results):
		super(PopupWindow, self).__init__()
		self.products = products 
		self.results = results

		# Load the ui file
		uic.loadUi(os.path.join('ui', 'popup.ui'), self) 

		# Add the 'select_location' function to the 'select_location_pushButton'
		self.select_location_pushButton.clicked.connect(self.select_location)

	def select_location(self):
		# Open dialog box
		self.file_path = QFileDialog.getExistingDirectory()
		# Run reate_spreadsheet function
		self.create_spreadsheet()

	def create_spreadsheet(self):
		# Create workbook
		wb = Workbook()

		for i in range(len(self.products)):
			# Crete product sheet
			wb.create_sheet(f'Product_{i+1}')
			# Set product sheet active
			wb.active = wb[f'Product_{i+1}']
			ws = wb.active
			# Add header to the sheet
			header = ['Product name', 'Pharmacy', 'Product price [PLN]', 'Product website']
			ws.append(header)
			# Style header
			for col in range(1, len(header) + 1):
				ws[get_column_letter(col) + '1'].font = Font(name='Calibri', bold=True)
			# Add results to the sheet
			for j in range(len(self.results[i])):
				if self.results[i][j] != {}:
					for key in self.results[i][j]:
						col_values = list(self.results[i][j][key].values())
						ws.append(col_values)
		# Remove default sheet
		wb.remove(wb['Sheet'])

		# Save the spreadsheet
		file_full_path = os.path.join(self.file_path, 'pharmacy_scrapper_comparison.xlsx')
		wb.save(file_full_path)

		# Open the spreadsheet
		# For Windows
		try:
			os.system(file_full_path)
		except:
			pass
		# For Linux
		try:
			subprocess.run(['open', file_full_path], check=True)
		except:
			pass


# Initialize the app
app = QApplication(sys.argv)
main_window = MainWindow()
main_window.setMinimumWidth(800)
main_window.setMinimumHeight(670)
main_window.show()
app.exec_()

